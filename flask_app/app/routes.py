from flask import Blueprint, jsonify, request, render_template, redirect, url_for
from .service import *
from .repository_mysql import *
from .repository_PostgreSQL import *
from openpyxl import Workbook

import os
import openpyxl

bp = Blueprint('/', __name__)

DB_TYPE = os.getenv('DB_TYPE', 'mysql')

if DB_TYPE == 'mysql':
    repository = BrandRepositoryMySQL()
elif DB_TYPE == 'postgresql':
    repository = BrandRepositoryPostgreSQL()
    
brandService = BrandService(repository=repository)
brandCodingsService = BrandCodingsService(repository=repository)
permutationsService = PermutationsService(repository=repository)
code_len = 0

def init_app(app):
    app.register_blueprint(bp)
    
@bp.route("/", methods =['GET'])
def index():
    return render_template("index.html")

@bp.route("/create_brand", methods = ["GET", "POST"])
def create_brand():
    if request.method == "POST":
        global code_len
        code_len = request.form.get('coding_len')
        brand_name = request.form.get('brand_name')
        brandService.create_brand(brand_name=brand_name)
        print("service function called")
        return redirect(url_for('/.create_coding', temp = code_len))
    elif request.method == "GET":
        return render_template("create_brand.html")
    
@bp.route("/create_coding/<int:temp>" , methods = ["GET", "POST"])
def create_coding(temp):
    if request.method == "POST":
        my_list = []
        global code_len
        code_len = temp
        instruction = ""
        for i in range(code_len):
            my_list.append([])
        
        for i in range(code_len):
            temp = request.form.get(str(i))
            instruction += request.form.get(str(i) + "-instruction")
            instruction += "-"
            holder = temp.split(" ")
            for j in range(len(holder)):
                my_list[i].append(holder[j])
            
        instruction = instruction.rstrip("-")
        brandCodingsService.create_coding(code_len, instruction)
        permutationsService.create_permutation(my_list=my_list)
        return redirect(url_for('/.index'))
    elif request.method == "GET":
        num = request.view_args['temp']
        return render_template("create_coding.html" , count = num)
    
@bp.route("/get_brands_instruction", methods = ["GET"])
def get_brands_instruction():
    res = brandCodingsService.get_brands_instruction()
    return render_template("brand_instructions.html", res = res)

@bp.route("/search_product/<string:code>", methods = ["GET"])
def search_product(code):
    print(code)
    print("========")
    res = permutationsService.search_product(code)
    print(res)
    print("BAAA")
    return render_template("search.html", res = res)

@bp.route("/delete_brand", methods = ["GET", "POST"])
def delete_brand():
    if request.method == "POST":
        givenID = request.form.get("id")
        brandService.delete_brand(givenID)
        return redirect(url_for("/.index"))
    elif request.method == "GET":
        res = brandService.get_all_brands()
        return render_template("delete_brand.html", res = res)
    
@bp.route("/update_brand_name", methods = ["GET", "POST"])
def update_brand_name():
    if request.method =="POST":
        givenID = request.form.get("id")
        newName = request.form.get("name")
        brandService.update_brand(givenID, newName)
        return redirect(url_for("/.index"))
    elif request.method =="GET":
        res = brandService.get_all_brands()
        return render_template("update_name.html", res = res)
    
@bp.route("/update_codes", methods = ["GET", "POST"])
def update_codes():
    if request.method == "POST":
        givenID = request.form.get("id")
        newCodeLen = request.form.get("code_len")
        brandCodingsService.update_coding(givenId=givenID) 
        return redirect(url_for("/.create_coding", temp = newCodeLen))
    elif request.method == "GET":
        res = brandService.get_all_brands()
        return render_template("update_code.html", res = res)
    
    
@bp.route("/excel")
def execl():
    data = openpyxl.load_workbook("app\\book1.xlsx")
    dataframe = data.active
    my_list = []
    for i in range(dataframe.max_row):
        my_list.append([])

    for i in range(1, dataframe.max_row+1):
        for j in range(1, dataframe.max_column+1):
            if dataframe.cell(row = i, column = j).value != None or j == 1:
                my_list[i-1].append(dataframe.cell(row= i, column =j).value)
                
    data.close()
    
    print(my_list)
    
    wb = Workbook()
    ws = wb.active
      
    for i in range(len(my_list)):
        if my_list[i][0] == 1:
            print("DONE!")
            res = brandCodingsService.get_brands_instruction()
            for i in res:
                temp = [i.brand_name, i.coding_instruction]
                ws.append(temp)
            wb.save("D:\\Prog\\Patronet\\Phase2\\flask_app\\app\\res.xlsx")
        elif my_list[i][0] == 2:
            givenID = my_list[i][1]
            brandService.delete_brand(givenID)
        elif my_list[i][0] == 3 or my_list[i][0] == 4:
            brand_name = ""
            givenID = -1
            new_code_len = my_list[i][2]
            temp_list = []
            instruction = ""
            for j in range(new_code_len):
                temp_list.append([])
            
            for j in range(new_code_len):
                instruction += my_list[i+j+1][1]
                instruction += "-"
                temp = my_list[i+j+1][2]
                holder = temp.split(" ")
                for k in range(len(holder)):
                    temp_list[j].append(holder[k])
                    
            instruction = instruction.rstrip("-") 
            if my_list[i][0] == 3:
                givenID = my_list[i][1]
                brandCodingsService.update_coding(givenId=givenID)
            else:
                brand_name = my_list[i][1]   
                brandService.create_brand(brand_name=brand_name)
 
            brandCodingsService.create_coding(new_code_len, instruction)
            permutationsService.create_permutation(my_list=temp_list)        
        else:
            continue
    return redirect(url_for("/.index"))
